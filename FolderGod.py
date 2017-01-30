#!/usr/bin/python

import sys
import os
import openpyxl


FolderList = openpyxl.load_workbook('List.xlsx')
Sheet = FolderList.get_sheet_by_name('Sheet1')
LastRow = int(Sheet.max_row)

cwd = os.getcwd()

for i in range(1,LastRow + 1):
    FolderName = str(Sheet.cell(row=i, column=5).value)
    if os.path.isdir(FolderName) == True:
        FolderName = FolderName + "(" + str(i) + ")"
        os.makedirs(cwd + '/' + FolderName)
    else:
        FolderName = FolderName.translate(None,"[/\?:*|<>]")
        os.makedirs(cwd + '/' + FolderName)